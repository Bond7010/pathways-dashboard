#!/usr/bin/env python3
"""
generate-dashboard.py
Reads pathway-record-master.xlsx (MY RECORD sheet) → bond-pathways-dashboard.html
Static HTML, no JS, FTH-compatible (no html/head/body wrapper).
"""
import openpyxl
from collections import OrderedDict
from html import escape
from datetime import date

XLSX = "pathway-record-master.xlsx"
OUT  = "bond-pathways-dashboard.html"

LEVEL_SORT = {
    'LEVEL ONE: Mastering Fundamentals':     0,
    'LEVEL TWO: Learning Your Style':        1,
    'LEVEL THREE: Increasing Knowledge':     2,
    'LEVEL FOUR: Building Skills':           3,
    'LEVEL FIVE: Demonstrating Expertise':   4,
}

CSS = """\
@import url('https://fonts.googleapis.com/css2?family=Montserrat:wght@400;600;700&family=Source+Sans+3:wght@400;600&display=swap');

*, *::before, *::after { box-sizing: border-box; margin: 0; padding: 0; }

:root {
  --blue:    #004165;
  --blue-lt: #0065a0;
  --maroon:  #772432;
  --purple:  #4A3468;
  --yellow:  #F2DF74;
  --amber:   #b86e00;
  --text:    #1a1a1a;
  --muted:   #666;
  --border:  #e0e0e0;
}

body {
  font-family: 'Source Sans 3', Arial, sans-serif;
  color: var(--text);
  background: #f5f5f5;
  font-size: 14px;
}

/* ── Header ────────────────────────────────────────────── */
header {
  background: var(--blue);
  color: #fff;
  padding: 0.75rem 1.5rem;
  display: flex;
  align-items: center;
  gap: 0.75rem;
}
.hdr-title {
  font-family: 'Montserrat', sans-serif;
  font-size: 1.1rem;
  font-weight: 700;
  letter-spacing: 0.02em;
  flex: 1;
}
.dtm-badge {
  background: var(--yellow);
  color: #000;
  font-family: 'Montserrat', sans-serif;
  font-size: 0.65rem;
  font-weight: 700;
  text-transform: uppercase;
  letter-spacing: 0.08em;
  padding: 0.2rem 0.5rem;
  border-radius: 3px;
}
.member-id {
  font-family: 'Montserrat', sans-serif;
  font-size: 0.75rem;
  opacity: 0.8;
}

/* ── Main / scroll ──────────────────────────────────────── */
main { padding: 0.75rem 1rem 1rem; overflow-x: hidden; }

#paths-scroll {
  display: flex;
  gap: 0.75rem;
  overflow-x: scroll;
  -webkit-overflow-scrolling: touch;
  padding-bottom: 1rem;
  width: 100%;
  min-width: 0;
}
#paths-scroll::-webkit-scrollbar { height: 8px; }
#paths-scroll::-webkit-scrollbar-track { background: #e0e0e0; border-radius: 4px; }
#paths-scroll::-webkit-scrollbar-thumb { background: var(--blue); border-radius: 4px; }
#paths-scroll::-webkit-scrollbar-thumb:hover { background: var(--maroon); }

/* ── Path column ────────────────────────────────────────── */
.path-col {
  flex: 0 0 260px;
  min-width: 260px;
  background: #fff;
  border-radius: 6px;
  box-shadow: 0 1px 4px rgba(0,0,0,.12);
  overflow: hidden;
  display: flex;
  flex-direction: column;
}
.path-hdr {
  padding: 0.6rem 0.75rem;
  color: #fff;
}
.path-name {
  font-family: 'Montserrat', sans-serif;
  font-size: 0.8rem;
  font-weight: 700;
  text-transform: uppercase;
  letter-spacing: 0.04em;
  margin-bottom: 0.3rem;
  line-height: 1.2;
}
.path-meta { display: flex; align-items: center; gap: 0.4rem; flex-wrap: wrap; }
.path-prog { font-size: 0.7rem; opacity: 0.85; margin-left: auto; }

/* ── Category badge ─────────────────────────────────────── */
.cat-badge {
  font-family: 'Montserrat', sans-serif;
  font-size: 0.58rem;
  font-weight: 700;
  text-transform: uppercase;
  letter-spacing: 0.06em;
  padding: 0.1rem 0.35rem;
  border-radius: 3px;
  border: 1px solid rgba(255,255,255,.5);
  color: #fff;
}

/* ── Status pills ───────────────────────────────────────── */
.pill {
  font-family: 'Montserrat', sans-serif;
  font-size: 0.6rem;
  font-weight: 700;
  text-transform: uppercase;
  letter-spacing: 0.06em;
  padding: 0.12rem 0.45rem;
  border-radius: 3px;
  white-space: nowrap;
}
.pill-complete { background: var(--maroon); color: #fff; }
.pill-inprog   { background: var(--yellow); color: #000; }
.pill-blank    { color: var(--muted); font-size: 0.8rem; background: transparent; }

/* ── Path body ──────────────────────────────────────────── */
.path-body { flex: 1; overflow-y: auto; }

/* ── Level sections ─────────────────────────────────────── */
.level-hdr {
  font-family: 'Montserrat', sans-serif;
  font-size: 0.63rem;
  font-weight: 700;
  text-transform: uppercase;
  letter-spacing: 0.05em;
  color: #fff;
  background: #888;
  padding: 0.2rem 0.6rem;
}
.level-complete .level-hdr { background: var(--maroon); }

/* ── Project rows ───────────────────────────────────────── */
.proj-row {
  display: flex;
  align-items: baseline;
  gap: 0.3rem;
  padding: 0.22rem 0.6rem;
  border-bottom: 1px solid var(--border);
  font-size: 0.76rem;
  line-height: 1.3;
}
.proj-row:last-child { border-bottom: none; }
.proj-name { flex: 1; }
.row-enh .proj-name { font-style: italic; color: var(--amber); font-size: 0.7rem; }

/* Elective badge */
.e-badge {
  flex: 0 0 auto;
  background: #e8e8e8;
  color: #555;
  font-family: 'Montserrat', sans-serif;
  font-size: 0.55rem;
  font-weight: 700;
  padding: 0.08rem 0.3rem;
  border-radius: 3px;
  letter-spacing: 0.03em;
}

/* ── Achievements ───────────────────────────────────────── */
.ach-section {
  margin-top: 1rem;
  padding: 0.75rem 1rem;
  background: #fff;
  border-radius: 6px;
  box-shadow: 0 1px 4px rgba(0,0,0,.10);
}
.ach-section h2 {
  font-family: 'Montserrat', sans-serif;
  font-size: 0.88rem;
  font-weight: 700;
  text-transform: uppercase;
  letter-spacing: 0.04em;
  color: var(--blue);
  margin-bottom: 0.5rem;
}
.ach-count {
  background: var(--maroon);
  color: #fff;
  border-radius: 10px;
  padding: 0.1rem 0.45rem;
  font-size: 0.72rem;
  vertical-align: middle;
}
.ach-pills { display: flex; flex-wrap: wrap; gap: 0.4rem; }
.ach-pill {
  font-family: 'Montserrat', sans-serif;
  font-size: 0.7rem;
  font-weight: 700;
  text-transform: uppercase;
  letter-spacing: 0.05em;
  padding: 0.25rem 0.6rem;
  border-radius: 3px;
  color: #fff;
}
.ach-current { background: var(--blue); }
.ach-legacy  { background: var(--maroon); }
.ach-vintage { background: var(--purple); }

/* ── Footer ─────────────────────────────────────────────── */
footer {
  margin-top: 1rem;
  padding: 0.75rem 1rem;
  font-size: 0.68rem;
  color: var(--muted);
  text-align: center;
  border-top: 1px solid var(--border);
  line-height: 1.6;
}

/* ── Print ──────────────────────────────────────────────── */
@media print {
  @page { size: A3 landscape; margin: 5mm; }
  body { zoom: 0.38; -webkit-print-color-adjust: exact; print-color-adjust: exact; }
  header { position: static !important; }
  #paths-scroll {
    overflow: visible !important;
    display: flex !important;
    flex-wrap: nowrap !important;
  }
  .path-col { break-inside: avoid; page-break-inside: avoid; }
  .ach-section { page-break-before: avoid; }
}"""


# ── Data loading ─────────────────────────────────────────────────────────────

def load():
    wb = openpyxl.load_workbook(XLSX, read_only=True)
    ws = wb[wb.sheetnames[2]]
    rows = []
    for row in ws.iter_rows(values_only=True):
        cells = list(row) + [None] * 8
        path, cat, level, project, kind, enh, status = cells[:7]
        if not path or not cat or not project:
            continue
        if path == 'PATH':  # header row
            continue
        if project in ('Current Path', 'Legacy Path', 'Vintage Path', 'Active Path'):
            continue
        rows.append(dict(
            path    = path.strip(),
            cat     = cat.strip().lower(),
            level   = level.strip() if level else '',
            project = project.strip(),
            kind    = kind.strip() if kind else '',
            is_enh  = (kind or '').strip() in ('Meeting Role', 'Ed Series'),
            status  = status.strip().upper() if status else None,
        ))
    return rows


def build_paths(rows):
    paths = OrderedDict()
    for r in rows:
        p = r['path']
        if p not in paths:
            paths[p] = dict(cat=r['cat'], levels=OrderedDict())
        lv = r['level']
        if lv not in paths[p]['levels']:
            paths[p]['levels'][lv] = []
        paths[p]['levels'][lv].append(r)
    return paths


# ── Status / progress ─────────────────────────────────────────────────────────

def level_complete(lv_rows):
    req = [r for r in lv_rows if r['kind'] == 'Required']
    elc = [r for r in lv_rows if r['kind'] == 'Elective']
    if not req:
        return False
    if any(r['status'] != 'COMPLETE' for r in req):
        return False
    if elc and not any(r['status'] == 'COMPLETE' for r in elc):
        return False
    return True


def path_status(pdata):
    has_any = any(r['status'] for lv in pdata['levels'].values() for r in lv)
    if not has_any:
        return 'NOT STARTED'
    if all(level_complete(lv) for lv in pdata['levels'].values()):
        return 'COMPLETE'
    return 'IN PROGRESS'


def progress(pdata):
    """(done, total) — Required rows + 1 elective slot per level that has electives."""
    done = total = 0
    for lv_rows in pdata['levels'].values():
        req = [r for r in lv_rows if r['kind'] == 'Required']
        elc = [r for r in lv_rows if r['kind'] == 'Elective']
        total += len(req)
        done  += sum(1 for r in req if r['status'] == 'COMPLETE')
        if elc:
            total += 1
            if any(r['status'] == 'COMPLETE' for r in elc):
                done += 1
    return done, total


# ── Sorting ───────────────────────────────────────────────────────────────────

def sorted_paths(paths):
    STATUS_RANK = {'IN PROGRESS': 0, 'COMPLETE': 1, 'NOT STARTED': 2}
    CAT_RANK    = {'current': 0, 'legacy': 1, 'vintage': 2}
    def key(item):
        pdata = item[1]
        return (STATUS_RANK.get(path_status(pdata), 3),
                CAT_RANK.get(pdata['cat'], 3))
    return sorted(paths.items(), key=key)


# ── HTML helpers ──────────────────────────────────────────────────────────────

def header_color(cat, status):
    if cat == 'current':
        return '#004165' if status == 'IN PROGRESS' else '#0065a0'
    if cat == 'legacy':
        return '#772432'
    if cat == 'vintage':
        return '#4A3468'
    return '#555'


def status_pill(status):
    if status == 'COMPLETE':
        return '<span class="pill pill-complete">COMPLETE</span>'
    if status == 'IN PROGRESS':
        return '<span class="pill pill-inprog">IN PROGRESS</span>'
    return '<span class="pill pill-blank">—</span>'


def cat_badge(cat):
    return f'<span class="cat-badge">{escape(cat.upper())}</span>'


LEVEL_SHORT = {'ONE': '1', 'TWO': '2', 'THREE': '3', 'FOUR': '4', 'FIVE': '5'}

def level_label(level):
    for word, num in LEVEL_SHORT.items():
        if f'LEVEL {word}' in level:
            return f'Level {num}'
    return level


def render_path_col(name, pdata):
    cat   = pdata['cat']
    st    = path_status(pdata)
    done, total = progress(pdata)
    hcol  = header_color(cat, st)

    if cat in ('current', 'legacy'):
        levels = sorted(pdata['levels'].items(),
                        key=lambda kv: LEVEL_SORT.get(kv[0], 99))
    else:
        levels = list(pdata['levels'].items())

    levels_html = ''
    for lv_name, lv_rows in levels:
        lv_done = level_complete(lv_rows)
        lv_cls  = ' level-complete' if lv_done else ''
        rows_html = ''
        for r in lv_rows:
            enh_cls = ' row-enh' if r['is_enh'] else ''
            e_badge = '<span class="e-badge">E</span> ' if r['kind'] == 'Elective' else ''
            rows_html += (
                f'<div class="proj-row{enh_cls}">'
                f'{e_badge}<span class="proj-name">{escape(r["project"])}</span>'
                f'{status_pill(r["status"])}'
                f'</div>\n'
            )
        levels_html += (
            f'<div class="level-section{lv_cls}">'
            f'<div class="level-hdr">{escape(level_label(lv_name))}</div>'
            f'{rows_html}'
            f'</div>\n'
        )

    return (
        f'<div class="path-col">'
        f'<div class="path-hdr" style="background:{hcol};">'
        f'<div class="path-name">{escape(name)}</div>'
        f'<div class="path-meta">'
        f'{cat_badge(cat)} {status_pill(st)}'
        f'<span class="path-prog">{done} / {total}</span>'
        f'</div></div>'
        f'<div class="path-body">{levels_html}</div>'
        f'</div>\n'
    )


def render(paths_sorted):
    today = date.today().isoformat()
    cols  = ''.join(render_path_col(n, d) for n, d in paths_sorted)

    achievements = [(n, d['cat']) for n, d in paths_sorted if path_status(d) == 'COMPLETE']
    ach_html = ''
    if achievements:
        pills = ''.join(
            f'<span class="ach-pill ach-{cat}">{escape(name)}</span>'
            for name, cat in achievements
        )
        ach_html = (
            f'<section class="ach-section">'
            f'<h2>Paths Completed <span class="ach-count">{len(achievements)}</span></h2>'
            f'<div class="ach-pills">{pills}</div>'
            f'</section>\n'
        )

    return (
        f'<style>\n{CSS}\n</style>\n\n'
        f'<header>\n'
        f'  <div class="hdr-title">Bond Wang — Pathways Dashboard</div>\n'
        f'  <span class="dtm-badge">DTM</span>\n'
        f'  <span class="member-id">EE4039</span>\n'
        f'</header>\n\n'
        f'<main>\n'
        f'  <div id="paths-scroll">\n{cols}  </div>\n'
        f'{ach_html}'
        f'</main>\n\n'
        f'<footer>\n'
        f'  <p>Toastmasters International is a registered trademark of Toastmasters International. '
        f'This page is not affiliated with or endorsed by Toastmasters International.</p>\n'
        f'  <p>Generated {today} from personal Pathways records.</p>\n'
        f'</footer>\n'
    )


if __name__ == '__main__':
    rows   = load()
    paths  = build_paths(rows)
    sorted_ = sorted_paths(paths)
    html   = render(sorted_)
    with open(OUT, 'w', encoding='utf-8') as f:
        f.write(html)
    print(f'Written: {OUT}')
    import os
    size = os.path.getsize(OUT)
    lines = html.count('\n') + 1
    print(f'Lines: {lines}  |  Size: {size:,} bytes')
